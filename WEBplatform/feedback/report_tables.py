import io
from tempfile import NamedTemporaryFile

from django.http import HttpResponse
from django.utils.text import slugify
from openpyxl.styles import Alignment

from groups.models import Student, LearningGroup
from openpyxl import Workbook


def create_student_report_table(student_id):
    student = Student.objects.filter(pk=student_id).first()
    response = None
    if student:
        student_feedback_lists = student.feedback_student_list.all()
        wb = Workbook()
        sheet = wb.active
        wrap_alignment = Alignment(wrap_text=True)
        sheet.column_dimensions['B'].width = 150
        sheet.cell(
            row=1,
            column=1,
            value=f'{student.lastname.title()} {student.name.title()} {student.patronymic.title()}'
        )
        sheet.cell(
            row=3,
            column=1,
            value='№'
        )
        sheet.cell(
            row=3,
            column=2,
            value="Обратная связь"
        )

        for index, student_list in enumerate(student_feedback_lists):
            sheet.cell(
                row=index + 4,
                column=1,
                value=index + 1
            )
            sheet.cell(
                row=index + 4,
                column=2,
                value=student_list.generated_report.text if student_list.generated_report.text != "" else "Не сгенерирована"
            ).alignment = wrap_alignment

        with io.BytesIO() as buffer:
            wb.save(buffer)
            stream = buffer.getvalue()
        response = HttpResponse(content=stream, content_type='application/ms-excel', )
        response['Content-Disposition'] = f'attachment; filename=report.xlsx'
    return response


def create_group_report_table(group_id):
    group = LearningGroup.objects.filter(pk=group_id).first()
    response = None
    if group:
        feedback_group_lists = group.feedback_group_list.all()
        wb = Workbook()
        sheet = wb.active
        wrap_alignment = Alignment(wrap_text=True)
        sheet.column_dimensions['B'].width = 150
        sheet.column_dimensions['A'].width = 50
        sheet.cell(
            row=1,
            column=1,
            value=f'{group.name.title()}'
        )

        for index_group_list, group_list in enumerate(feedback_group_lists):
            feedback_student_lists = group_list.feedback_student_list.all()
            date = group_list.feedback_list.date
            row = 3 + index_group_list * (len(feedback_student_lists) + 4)
            sheet.cell(
                row=row,
                column=1,
                value=f'Обратная связь от:   {date.strftime("%d.%m.%y")}'
            )
            sheet.cell(
                row=row + 1,
                column=1,
                value='ФИО'
            )
            sheet.cell(
                row=row + 1,
                column=2,
                value="Обратная связь"
            )
            for index_student_list, student_list in enumerate(feedback_student_lists):
                student = student_list.student
                sheet.cell(
                    row=row + 2 + index_student_list,
                    column=1,
                    value=f'{student.lastname.title()} {student.name.title()} {student.patronymic.title()}'
                )
                sheet.cell(
                    row=row + 2 + index_student_list,
                    column=2,
                    value=student_list.generated_report.text if student_list.generated_report.text != "" else "Не сгенерирована"
                ).alignment = wrap_alignment
        with io.BytesIO() as buffer:
            wb.save(buffer)
            stream = buffer.getvalue()
        response = HttpResponse(content=stream, content_type='application/ms-excel', )
        response['Content-Disposition'] = f'attachment; filename=report.xlsx'
    return response


def create_full_report_table():
    pass
